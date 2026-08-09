"""Extensible tabular ingestion with explicit source-to-domain mapping."""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from typing import Mapping

from .triage import OPTIONAL_COLUMNS, REQUIRED_COLUMNS, ValidationResult, validate_rows


@dataclass(frozen=True)
class RawTable:
    headers: tuple[str, ...]
    rows: tuple[dict[str, str], ...]
    format: str


class IngestionError(ValueError):
    """A user-correctable file or mapping problem."""


ALIASES = {
    "ticket_id": ("id", "ticket", "ticket number", "case id", "issue id"),
    "created_at": ("created", "opened at", "opened_at", "created date", "created timestamp"),
    "status": ("state", "ticket status", "workflow status"),
    "assignee": ("owner", "assigned to", "assigned_to", "agent", "analyst"),
    "team": ("group", "queue", "department", "support team"),
    "subject": ("title", "summary", "description", "issue", "request"),
    "priority": ("severity", "urgency", "ticket priority"),
    "last_updated_at": ("updated", "updated at", "updated_at", "modified at"),
}


def _header_key(value: object) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def _headers_from_row(values: list[object]) -> tuple[str, ...]:
    headers = tuple(str(value or "").strip() for value in values)
    if not headers or any(not header for header in headers):
        raise IngestionError("The first row must contain non-empty column names.")
    keys = [_header_key(header) for header in headers]
    if len(keys) != len(set(keys)):
        raise IngestionError("The file has duplicate column names. Rename duplicates before importing.")
    return headers


def _rows_from_matrix(headers: tuple[str, ...], matrix: list[list[object]]) -> tuple[dict[str, str], ...]:
    rows = []
    for values in matrix:
        padded = list(values) + [""] * max(0, len(headers) - len(values))
        rows.append(
            {
                header: "" if padded[index] is None else str(padded[index]).strip()
                for index, header in enumerate(headers)
            }
        )
    return tuple(rows)


def read_table(content: bytes, filename: str) -> RawTable:
    """Read CSV/TSV or XLSX into a neutral table before mapping."""

    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else "csv"
    if suffix in {"xlsx", "xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise IngestionError("XLSX support needs openpyxl; install the pinned project dependencies.") from exc
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            iterator = sheet.iter_rows(values_only=True)
            header_values = list(next(iterator))
            headers = _headers_from_row(header_values)
            rows = _rows_from_matrix(headers, [list(values) for values in iterator])
        except (StopIteration, TypeError, ValueError) as exc:
            raise IngestionError(f"Could not read XLSX file: {exc}") from exc
        return RawTable(headers, rows, "xlsx")

    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise IngestionError("CSV/TSV files must be UTF-8 encoded.") from exc
    if not text.strip():
        raise IngestionError("The uploaded file is empty.")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    try:
        headers = _headers_from_row(list(next(reader)))
    except StopIteration as exc:
        raise IngestionError("The uploaded file has no header row.") from exc
    rows = _rows_from_matrix(headers, [list(values) for values in reader])
    return RawTable(headers, rows, "csv" if dialect.delimiter == "," else "tsv")


def suggest_mapping(headers: tuple[str, ...]) -> dict[str, str | None]:
    """Suggest a mapping without hiding unmapped required fields."""

    by_key = {_header_key(header): header for header in headers}
    mapping: dict[str, str | None] = {}
    for canonical in (*REQUIRED_COLUMNS, *OPTIONAL_COLUMNS):
        candidates = (canonical, *ALIASES.get(canonical, ()))
        mapping[canonical] = next(
            (by_key[_header_key(candidate)] for candidate in candidates if _header_key(candidate) in by_key), None
        )
    return mapping


def map_rows(table: RawTable, mapping: Mapping[str, str | None]) -> list[dict[str, str]]:
    """Project source columns into the stable domain schema."""

    return [
        {
            canonical: row.get(source, "") if source else ""
            for canonical in (*REQUIRED_COLUMNS, *OPTIONAL_COLUMNS)
            for source in (mapping.get(canonical),)
        }
        for row in table.rows
    ]


def validate_table(table: RawTable, mapping: Mapping[str, str | None]) -> ValidationResult:
    mapped_rows = map_rows(table, mapping)
    return validate_rows(mapped_rows, (*REQUIRED_COLUMNS, *OPTIONAL_COLUMNS))
